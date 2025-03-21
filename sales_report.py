import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# File paths
input_file = "C:\\Excel_Projects\\Sales_Report_Automation\\sales_data.xlsx"
output_dir = "C:\\Excel_Projects\\Sales_Report_Automation\\output"
output_file = os.path.join(output_dir, "sales_report.xlsx")

# Ensure output directory exists
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Load Excel Data
df = pd.read_excel(input_file)

# Check and Rename Columns to match script expectations
column_mapping = {
    'quantity': 'Quantity',
    'unit_price': 'Unit Price',
    'total_price': 'Total Sales',
    'pizza_category': 'Category',
    'pizza_name': 'Product'
}
df.rename(columns=column_mapping, inplace=True)

# Calculate Total Sales if not in the file
if 'Total Sales' not in df.columns:
    df["Total Sales"] = df["Quantity"] * df["Unit Price"]

# Sales Summary by Category
category_summary = df.groupby("Category")["Total Sales"].sum().reset_index()

# Top 5 Selling Products
top_products = df.groupby("Product")["Total Sales"].sum().nlargest(5).reset_index()

# Generate Charts
plt.figure(figsize=(6, 4))
plt.bar(category_summary["Category"], category_summary["Total Sales"], color="skyblue")
plt.xlabel("Category")
plt.ylabel("Total Sales")
plt.title("Sales by Category")
plt.xticks(rotation=45)
plt.tight_layout()

chart_path = os.path.join(output_dir, "sales_chart.png")
plt.savefig(chart_path)
plt.close()

# Write Processed Data & Summary to a New Excel File
wb = load_workbook(input_file)
ws = wb.active

# Append Data
ws.append([])
ws.append(["Sales Summary"])
for r in dataframe_to_rows(category_summary, index=False, header=True):
    ws.append(r)

# Format Header
header_font = Font(bold=True)
header_fill = PatternFill(start_color="FFFF00", fill_type="solid")
for cell in ws["A1:Z1"][0]:  
    cell.font = header_font
    cell.fill = header_fill

# Save Excel Report
wb.save(output_file)
print(f"✅ Sales report generated successfully: {output_file}")
